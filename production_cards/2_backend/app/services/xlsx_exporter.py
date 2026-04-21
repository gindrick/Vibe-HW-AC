from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def _thin() -> Side:
    return Side(style="thin")


def _border(left=False, right=False, top=False, bottom=False) -> Border:
    return Border(
        left=_thin() if left else Side(),
        right=_thin() if right else Side(),
        top=_thin() if top else Side(),
        bottom=_thin() if bottom else Side(),
    )


def _all_borders() -> Border:
    return Border(left=_thin(), right=_thin(), top=_thin(), bottom=_thin())


def _header_fill() -> PatternFill:
    return PatternFill("solid", fgColor="D9D9D9")


def _apply(ws, row: int, col: int, value: Any, bold=False, italic=False, size=10,
           fill=None, border=None, wrap=False, align_h="left", align_v="center") -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Calibri", bold=bold, italic=italic, size=size)
    cell.alignment = Alignment(horizontal=align_h, vertical=align_v, wrap_text=wrap)
    if fill:
        cell.fill = fill
    if border:
        cell.border = border


def _section_header(ws, row: int, col_start: int, col_end: int, text: str) -> None:
    """Gray header label spanning col_start:col_end."""
    _apply(ws, row, col_start, text, bold=True, size=9,
           fill=_header_fill(), border=_all_borders(), align_h="center")
    if col_end > col_start:
        ws.merge_cells(
            start_row=row, start_column=col_start,
            end_row=row, end_column=col_end,
        )


def _section_value(ws, row: int, col_start: int, col_end: int, text: str) -> None:
    """Value cell spanning col_start:col_end."""
    _apply(ws, row, col_start, text, size=10, border=_all_borders())
    if col_end > col_start:
        ws.merge_cells(
            start_row=row, start_column=col_start,
            end_row=row, end_column=col_end,
        )


# Column layout (1-based):
# A=1  B=2  C=3  D=4  E=5  F=6  G=7
# Three-column sections: A-B | C-D | E-G
# Measurements (5):      A   | B   | C   | D   | E-G
# Parameters:            A(#)| B-D (Parametr) | E-F (Hodnota) | G (Jedn.)

_C1, _C2 = 1, 2   # left third
_C3, _C4 = 3, 4   # middle third
_C5, _C7 = 5, 7   # right third (and last col)

COL_WIDTHS = [5, 20, 14, 10, 18, 12, 10]  # A-G


def build_xlsx(card: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Nastavovací karta"

    for i, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    parameters: list[dict] = card.get("parameters") or []
    r = 1

    # ── Title ──────────────────────────────────────────────────────────────
    ws.row_dimensions[r].height = 22
    _apply(ws, r, 1, card.get("title") or "NASTAVOVACÍ KARTA EXTRUZNÍ LINKY",
           bold=True, size=12, align_h="center", border=_all_borders())
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    r += 1

    # ── Datum ───────────────────────────────────────────────────────────────
    _section_header(ws, r, 1, 7, "VÝROBA ABS – DATUM")
    r += 1
    _section_value(ws, r, 1, 7, card.get("date") or "")
    r += 1

    # ── Číslo linky | Směna | Obsluha ───────────────────────────────────────
    _section_header(ws, r, _C1, _C2, "ČÍSLO VÝROBNÍ LINKY")
    _section_header(ws, r, _C3, _C4, "SMĚNA")
    _section_header(ws, r, _C5, _C7, "OBSLUHA")
    r += 1
    _section_value(ws, r, _C1, _C2, card.get("line_number") or "")
    _section_value(ws, r, _C3, _C4, card.get("shift") or "")
    _section_value(ws, r, _C5, _C7, card.get("operator") or "")
    r += 1

    # ── Nástroj | Vyráběný rozměr | Povrchová úprava ────────────────────────
    _section_header(ws, r, _C1, _C2, "NÁSTROJ")
    _section_header(ws, r, _C3, _C4, "VYRÁBĚNÝ ROZMĚR")
    _section_header(ws, r, _C5, _C7, "POVRCHOVÁ ÚPRAVA")
    r += 1
    _section_value(ws, r, _C1, _C2, card.get("tool") or "")
    _section_value(ws, r, _C3, _C4, card.get("produced_dimension") or "")
    _section_value(ws, r, _C5, _C7, card.get("surface_treatment") or "")
    r += 1

    # ── Číslo artiklu | Materiál/Granulát | Lakování ────────────────────────
    _section_header(ws, r, _C1, _C2, "ČÍSLO ARTIKLU")
    _section_header(ws, r, _C3, _C4, "MATERIÁL/GRANULÁT")
    _section_header(ws, r, _C5, _C7, "LAKOVÁNÍ (DRUH, POMĚR)")
    r += 1
    _section_value(ws, r, _C1, _C2, card.get("article_number") or "")
    _section_value(ws, r, _C3, _C4, card.get("material_granulate") or "")
    _section_value(ws, r, _C5, _C7, card.get("coating") or "")
    r += 1

    # ── Obsluha naměřeno – section title ────────────────────────────────────
    _apply(ws, r, 1, "OBSLUHA NAMĚŘENO", bold=True, size=9,
           fill=_header_fill(), border=_all_borders(), align_h="center")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    r += 1

    # ── Measurement headers ──────────────────────────────────────────────────
    _section_header(ws, r, 1, 2, "TLOUŠŤKA")
    _section_header(ws, r, 3, 3, "ŠÍŘKA")
    _section_header(ws, r, 4, 4, "U PROFIL")
    _section_header(ws, r, 5, 6, "POVRCH")
    _section_header(ws, r, 7, 7, "LESK *")
    r += 1

    # ── Measurement values ───────────────────────────────────────────────────
    _section_value(ws, r, 1, 2, card.get("thickness") or "")
    _section_value(ws, r, 3, 3, card.get("width") or "")
    _section_value(ws, r, 4, 4, card.get("u_profile") or "")
    _section_value(ws, r, 5, 6, card.get("surface") or "")
    _section_value(ws, r, 7, 7, card.get("gloss") or "")
    r += 1

    # ── Parameters table header ──────────────────────────────────────────────
    _apply(ws, r, 1, "#", bold=True, size=9,
           fill=_header_fill(), border=_all_borders(), align_h="center")
    _apply(ws, r, 2, "PARAMETR", bold=True, size=9,
           fill=_header_fill(), border=_all_borders())
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    _apply(ws, r, 5, "HODNOTA", bold=True, size=9,
           fill=_header_fill(), border=_all_borders())
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
    _apply(ws, r, 7, "JEDN.", bold=True, size=9,
           fill=_header_fill(), border=_all_borders(), align_h="center")
    r += 1

    # ── Parameter rows ───────────────────────────────────────────────────────
    for param in parameters:
        ws.row_dimensions[r].height = 15
        _apply(ws, r, 1, param.get("number"), border=_all_borders(), align_h="center")
        _apply(ws, r, 2, param.get("name") or "", border=_all_borders())
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        _apply(ws, r, 5, param.get("value") or "", border=_all_borders())
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
        _apply(ws, r, 7, param.get("unit") or "", border=_all_borders(), align_h="center")
        r += 1

    # ── Poznámky ─────────────────────────────────────────────────────────────
    _apply(ws, r, 1, "POZNÁMKY", bold=True, size=9,
           fill=_header_fill(), border=_all_borders(), align_h="center")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    r += 1
    _apply(ws, r, 1, card.get("notes") or "",
           wrap=True, border=_all_borders(), align_v="top")
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=7)
    ws.row_dimensions[r].height = 50
    r += 3

    # ── Footer ───────────────────────────────────────────────────────────────
    r += 1
    zpracoval = card.get("footer_processed_by") or ""
    schvalil = card.get("footer_approved_by") or ""
    _apply(ws, r, 1, f"Zpracoval: {zpracoval}", italic=True, size=9)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    _apply(ws, r, 3, "stránka 1 z 1", size=9, align_h="center")
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    _apply(ws, r, 6, f"Schválil: {schvalil}", italic=True, size=9)
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
